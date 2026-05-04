import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
import copy
from datetime import datetime

EXCEL_PATH = '/home/user/ugm_ay/2026_9-32 ÜGM Genelge_Ek_ÜGM  Başvuruları Takip Listesi-2026__9-32.xlsx'

# Her başvuru için veriler: (ay, yıl, takip_no, basvuru_tarihi, sanayici_unvani,
#  vergi_no, mersis_no, adres, durum, nihai_urunler)
# nihai_urunler: [(nihai_urun_adi, [(girdi_adi, miktar, birim, gtip)])]

BASVU_DATA = [
    # --- OCAK ---
    {
        'ay': 'OCAK', 'yil': 2026,
        'takip_no': '2026/00096',
        'basvuru_tarihi': '09.01.2026',
        'sanayici_unvani': 'ALİNDAİR SOĞUTMA SİSTEMLERİ SANAYİ VE TİCARET ANONİM ŞİRKETİ',
        'vergi_no': '0530524182',
        'mersis_no': '0053052418200044',
        'adres': 'PINARKENT MAH. 210 SK. NO:3 PAMUKKALE / DENİZLİ',
        'durum': 'OLUMLU',
        'nihai_urunler': [
            ('EGZOZ FANI', [
                ('AC MOTOR', 39000, 'Adet', '8501.52.20.90.19'),
                ('AC MOTOR', 5000, 'Adet', '8501.51.00.90.00'),
            ]),
            ('SANAYİ TİPİ EVAPORATİF SOĞUTUCU', [
                ('UV LAMBA', 10000, 'Adet', '8539.49.00.00.00'),
                ('DC MOTOR', 20000, 'Adet', '8501.32.00.90.00'),
                ('TOZ FİLTRESİ', 160000, 'Adet', '8421.39.25.90.00'),
                ('ANAKART', 39000, 'Adet', '8537.10.91.00.00'),
                ('TOZ FİLTRESİ', 42000, 'Metre', '8421.39.25.90.00'),
                ('AC MOTOR', 19000, 'Adet', '8501.52.20.90.19'),
                ('DUVAR KUMANDASI', 39000, 'Adet', '8537.10.91.00.00'),
                ('SU POMPASI', 39000, 'Adet', '8413.81.00.00.00'),
            ]),
            ('EVSEL TİP EVAPORATİF SOĞUTUCU', [
                ('SWING MOTOR', 185000, 'Adet', '8501.40.20.29.00'),
                ('TOZ FİLTRESİ', 475000, 'Adet', '8421.39.25.90.00'),
                ('AC MOTOR', 155000, 'Adet', '8501.40.20.29.00'),
                ('SU POMPASI', 155000, 'Adet', '8413.81.00.00.00'),
                ('ANAKART', 310000, 'Adet', '8537.10.91.00.00'),
            ]),
        ],
    },
    {
        'ay': 'OCAK', 'yil': 2026,
        'takip_no': '2026/00186',
        'basvuru_tarihi': '12.01.2026',
        'sanayici_unvani': 'PANDA MOTOR TEKNOLOJİ SANAYİ TİCARET LİMİTED ŞİRKETİ',
        'vergi_no': '7210934529',
        'mersis_no': '0721093452900001',
        'adres': 'HACIEYÜPLÜ MAH. 3105. SK. NO: 3 İÇ KAPI NO: Z2 MERKEZEFENDİ / DENİZLİ',
        'durum': 'OLUMSUZ',
        'nihai_urunler': [
            ('DÖRT TEKER ELEKTRİKLİ ARAÇ', [
                ('KONVERTÖR', 8100, 'Adet', '8504.40.83.90.19'),
            ]),
            ('ELEKTRİKLİ MOPED', [
                ('KONVERTÖR', 36000, 'Adet', '8504.40.83.90.19'),
            ]),
            ('ELEKTRİKLİ ÜÇ TEKER (YOLCU VE YÜK TAŞIMA)', [
                ('KONVERTÖR', 36000, 'Adet', '8504.40.83.90.19'),
            ]),
        ],
    },
    {
        'ay': 'OCAK', 'yil': 2026,
        'takip_no': '2026/00365',
        'basvuru_tarihi': '12.01.2026',
        'sanayici_unvani': 'ST GRUP ENDÜSTRİYEL ÜRÜNLER BİLGİSAYAR SANAYİ VE TİCARET LİMİTED ŞİRKETİ HONAZ ŞUBESİ',
        'vergi_no': '7810409269',
        'mersis_no': '0781040926900039',
        'adres': 'ORGANİZE SANAYİ BÖLGESİ VALİ MÜNİR GÜNEY CADDESİ NO:6 HONAZ/DENİZLİ',
        'durum': 'OLUMLU',
        'nihai_urunler': [
            ('MOTOSİKLET', [
                ('CONVERTÖR', 30400, 'Adet', '8504.40.95.90.19'),
                ('İNVERTÖR', 30400, 'Adet', '8504.40.84.90.00'),
            ]),
            ('ELEKTRİKLİ BİSİKLET', [
                ('CONVERTÖR', 24000, 'Adet', '8504.40.95.90.19'),
                ('İNVERTÖR', 24000, 'Adet', '8504.40.84.90.00'),
            ]),
            ('ELEKTRİKLİ BİSİKLET (E-TRİCYCLE)', [
                ('CONVERTÖR', 50000, 'Adet', '8504.40.95.90.19'),
                ('İNVERTÖR', 50000, 'Adet', '8504.40.84.90.00'),
            ]),
            ('MOPED MOTOSİKLET (ELEKTRİKLİ MOPED)', [
                ('İNVERTÖR', 76000, 'Adet', '8504.40.84.90.00'),
                ('CONVERTÖR', 76000, 'Adet', '8504.40.95.90.19'),
            ]),
        ],
    },
    {
        'ay': 'OCAK', 'yil': 2026,
        'takip_no': '2026/00846',
        'basvuru_tarihi': '19.01.2026',
        'sanayici_unvani': 'PANDA MOTOR TEKNOLOJİ SANAYİ TİCARET LİMİTED ŞİRKETİ',
        'vergi_no': '7210934529',
        'mersis_no': '0721093452900001',
        'adres': 'HACIEYÜPLÜ MAH. 3105. SK. NO: 3 İÇ KAPI NO: Z2 MERKEZEFENDİ / DENİZLİ',
        'durum': 'OLUMLU',
        'nihai_urunler': [
            ('DÖRT TEKER ELEKTRİKLİ ARAÇ', [
                ('KONVERTÖR', 8100, 'Adet', '8504.40.83.90.19'),
            ]),
            ('ELEKTRİKLİ MOPED', [
                ('KONVERTÖR', 36000, 'Adet', '8504.40.83.90.19'),
            ]),
            ('ELEKTRİKLİ ÜÇ TEKER (YOLCU VE YÜK TAŞIMA)', [
                ('KONVERTÖR', 36000, 'Adet', '8504.40.83.90.19'),
            ]),
        ],
    },
    # --- ŞUBAT ---
    {
        'ay': 'ŞUBAT', 'yil': 2026,
        'takip_no': '2026/01440',
        'basvuru_tarihi': '05.02.2026',
        'sanayici_unvani': 'HÜNKAR ECZA VE MEDİKAL SANAYİ VE TİCARET LİMİTED ŞİRKETİ',
        'vergi_no': '4640051936',
        'mersis_no': '0464005193600019',
        'adres': 'AKÇEŞME MAHALLESİ 2019 SOKAK NO:5 MERKEZEFENDİ/DENİZLİ',
        'durum': 'OLUMLU',
        'nihai_urunler': [
            ('GİYİLEBİLİR MESH NEBULİZATÖR', [
                ('Elektronik kart (ELEKTRONİK KART PCB KART)', 109000, 'Adet', '8537.10.91.00.00'),
            ]),
            ('NEBÜLİZATÖR', [
                ('Tetik anahtarı (SWİTEK (ANAHTAR)(2,5-16 AMPER))', 694800, 'Adet', '8536.69.90.00.18'),
                ('Solunum cihazı kompresörü (DÖNER HAREKETLİ KOMPRESÖR)', 694800, 'Adet', '8414.80.73.90.00'),
            ]),
            ('DİJİTAL TANSİYON ALETİ ÜRETİMİ', [
                ('Elektronik kart (ELEKTRONİK ANA PCB KART)', 625000, 'Adet', '8537.10.91.00.00'),
            ]),
            ('ELEKTRİKLİ GÖĞÜS POMPASI', [
                ('Adaptör (AC/DC ADAPTÖR (6V 1A))', 41000, 'Adet', '8504.40.83.90.19'),
                ('Elektronik kart (ELEKTRONİK ANA PCB KART)', 31000, 'Adet', '8537.10.91.00.00'),
                ('Vakum pompası (DC 6V MIKRO HAVA VAKUM POMPASI)', 41000, 'Adet', '8501.31.00.90.11'),
            ]),
            ('ODA DERECESİ', [
                ('Elektronik kart (ELEKTRONİK KART PCB KART)', 1000, 'Adet', '8537.10.91.00.00'),
            ]),
            ('TAŞINABİLİR MESH NEBULİZATÖR', [
                ('Elektronik kart (ELEKTRONİK ANA PCB KART)', 72000, 'Adet', '8537.10.91.00.00'),
            ]),
            ('DİJİTAL TEMASSIZ ATEŞ ÖLÇER', [
                ('Elektronik kart (ELEKTRONİK KART PCB KART)', 200000, 'Adet', '8537.10.91.00.00'),
            ]),
        ],
    },
    # --- MART ---
    {
        'ay': 'MART', 'yil': 2026,
        'takip_no': '2026/01767',
        'basvuru_tarihi': '25.02.2026',
        'sanayici_unvani': 'ALİNDAİR SOĞUTMA SİSTEMLERİ SANAYİ VE TİCARET ANONİM ŞİRKETİ',
        'vergi_no': '0530524182',
        'mersis_no': '0053052418200044',
        'adres': 'PINARKENT MAH. 210 SK. NO:3 PAMUKKALE / DENİZLİ',
        'durum': 'OLUMLU',
        'nihai_urunler': [
            ('EGZOZ FAN', [
                ('AC MOTOR', 24000, 'Adet', '8501.51.00.90.00'),
                ('AC MOTOR', 3000, 'Adet', '8501.40.80.11.00'),
                ('AC MOTOR', 1500, 'Adet', '8501.40.20.29.00'),
                ('AC MOTOR', 3000, 'Adet', '8501.40.20.29.00'),
                ('AC MOTOR', 14000, 'Adet', '8501.52.20.90.19'),
            ]),
            ('EVSEL TİP EVAPORATİF SOĞUTUCU', [
                ('SU POMPASI', 77500, 'Adet', '8413.70.21.90.00'),
                ('TOZ FİLTRESİ', 475, 'Adet', '8421.39.25.90.00'),
                ('SWING MOTOR', 185000, 'Adet', '8501.40.20.29.00'),
                ('ANAKART', 155000, 'Adet', '8537.10.91.00.00'),
                ('SU POMPASI', 77500, 'Adet', '8413.81.00.00.00'),
                ('PLASTİK FİLE', 280250, 'Adet', '8421.39.25.90.00'),
                ('AC MOTOR', 155000, 'Adet', '8501.40.20.29.00'),
                ('SWING MOTOR', 185000, 'Adet', '8501.40.20.29.00'),
                ('ANAKART', 155000, 'Adet', '8504.40.83.90.19'),
            ]),
            ('SANAYİ TİPİ EVAPORATİF SOĞUTUCU', [
                ('DC MOTOR', 12000, 'Adet', '8501.32.00.90.00'),
                ('ANAKART', 19500, 'Adet', '8504.40.83.90.19'),
                ('DUVAR KUMANDASI', 39000, 'Adet', '8537.10.91.00.00'),
                ('AC MOTOR', 19000, 'Adet', '8501.52.20.90.19'),
                ('DC MOTOR', 8000, 'Adet', '8501.31.00.90.11'),
                ('TOZ FİLTRESİ', 160000, 'Adet', '8421.39.25.90.00'),
                ('UV LAMBA', 10000, 'Adet', '8539.49.00.00.00'),
                ('ANAKART', 19500, 'Adet', '8537.10.91.00.00'),
                ('SU POMPASI', 19500, 'Adet', '8413.70.21.90.00'),
                ('TOZ FİLTRESİ', 42000, 'Metre', '8421.39.25.90.00'),
                ('SU POMPASI', 19500, 'Adet', '8413.81.00.00.00'),
            ]),
        ],
    },
    {
        'ay': 'MART', 'yil': 2026,
        'takip_no': '2026/01785',
        'basvuru_tarihi': '26.02.2026',
        'sanayici_unvani': 'EGEVİZYON ELEKTRONİK MAKİNA SANAYİ VE TİCARET ANONİM ŞİRKETİ',
        'vergi_no': '3250866772',
        'mersis_no': '0325086677200001',
        'adres': 'IRLIGANLI MAH. ATATÜRK CAD. NO: 6 İÇ KAPI NO: 1 PAMUKKALE / DENİZLİ',
        'durum': 'OLUMSUZ',
        'nihai_urunler': [
            ('Evaporatif Soğutucu', [
                ('Kontrol Kartı', 4500, 'Adet', '8537.10.91.00.00'),
                ('İnvertör Kontrol Sistemi', 4500, 'Adet', '8504.40.87.90.11'),
                ('Hava Serinletici Gövde', 4500, 'Adet', '8509.80.00.00.00'),
                ('Su Pompası', 9000, 'Adet', '8413.70.21.90.00'),
                ('Filtre', 54000, 'Adet', '8421.39.25.90.00'),
                ('Motor', 9000, 'Adet', '8501.40.20.29.00'),
            ]),
            ('Fanlı Isıtıcı', [
                ('Fan', 1500, 'Adet', '8414.59.95.90.00'),
            ]),
            ('İnfrared Isıtıcı', [
                ('Kontrol Kartı', 7500, 'Adet', '8537.10.91.00.00'),
                ('İnfrared Lamba', 7500, 'Adet', '8539.49.00.00.00'),
            ]),
            ('Radyal-Aksiyal Fan', [
                ('Motor', 3000, 'Adet', '8501.52.20.90.19'),
            ]),
        ],
    },
    {
        'ay': 'MART', 'yil': 2026,
        'takip_no': '2026/02014',
        'basvuru_tarihi': '17.03.2026',
        'sanayici_unvani': 'ALSANCAK İNŞAAT TAAHHÜT MÜTEAHHİTLİK GAYRİMENKUL TARIM HAYVANCILIK OTOMOTİV HAFRİYAT NAKLİYAT TEKSTİL GIDA TURİZM DANIŞMANLIK SANAYİ VE TİCARET LİMİTED ŞİRKETİ',
        'vergi_no': '0590645661',
        'mersis_no': '0059064566100001',
        'adres': 'DENİZLİ OSB MAH. FAHRİ KARACA CAD. FABER MERMER BLOK NO: 17 HONAZ / DENİZLİ',
        'durum': 'OLUMLU',
        'nihai_urunler': [
            ('Egzoz Fanı', [
                ('Motor', 6375, 'Adet', '8501.52.20.90.11'),
            ]),
            ('Evaporatif Soğutucu Evsel Tipi', [
                ('Anakart', 1600, 'Adet', '8537.10.91.00.00'),
                ('Su Pompası', 1600, 'Adet', '8413.81.00.00.00'),
                ('Motor', 1600, 'Adet', '8501.40.20.29.00'),
            ]),
            ('Evaporatif Soğutucu Sanayi Tipi', [
                ('Anakart', 3200, 'Adet', '8537.10.91.00.00'),
                ('Motor', 3200, 'Adet', '8501.52.20.90.11'),
                ('Duvar Kumandası', 3200, 'Adet', '8537.10.91.00.00'),
                ('Su Pompası', 3200, 'Adet', '8413.81.00.00.00'),
            ]),
            ('Sirkülasyon Fanı', [
                ('Motor', 6375, 'Adet', '8501.51.00.90.00'),
            ]),
        ],
    },
]

# Kalan aylar (veri yok)
REMAINING_MONTHS = ['NİSAN', 'MAYIS', 'HAZİRAN', 'TEMMUZ', 'AĞUSTOS', 'EYLÜL', 'EKİM', 'KASIM', 'ARALIK']


def make_thin_border():
    thin = Side(style='thin')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def write_data_row(ws, row_num, ay, yil, basvuru, nihai_urun_adi, girdi):
    """Bir satır yaz"""
    b = basvuru
    girdi_adi, miktar, birim, gtip = girdi

    data = {
        1: 'DENİZLİ',       # A: İl Müdürlüğü
        2: ay,               # B: Ay
        3: yil,              # C: Yıl
        4: b['takip_no'],    # D: Başvuru Takip No
        5: b['basvuru_tarihi'],  # E: Başvuru Tarihi
        6: None,             # F: Muafiyet Yazısı Tarihi (PDF'de yok)
        7: None,             # G: Muafiyet Yazısı Sayısı (PDF'de yok)
        8: b['sanayici_unvani'],  # H: Sanayici Ünvanı
        9: b['vergi_no'],    # I: Sanayici Vergi No
        10: b['mersis_no'],  # J: Sanayici Mersis No
        11: b['adres'],      # K: Sanayici Adresi
        12: None,            # L: Tedarikçi Ünvanı
        13: None,            # M: Tedarikçi Vergi No
        14: None,            # N: Tedarikçi Adresi
        15: nihai_urun_adi,  # O: Nihai Ürün Adı
        16: gtip,            # P: Girdi GTIP
        17: girdi_adi,       # Q: Girdi Ürün Adı
        18: miktar,          # R: Kapasite Raporundaki Miktar
        19: miktar,          # S: ÜGM Yazısı kapsamında kullanılabilecek miktar (= rapor miktarı)
        20: birim,           # T: Miktar Birimi
        21: 'X' if b['durum'] == 'OLUMLU' else None,   # U: Olumlu
        22: 'X' if b['durum'] == 'OLUMSUZ' else None,  # V: Olumsuz
        23: None,            # W: Açıklama
        24: 'Evet',          # X: Bakanlığımız Sorumluluğunda mı?
        25: None,            # Y: TSE ÜGM ÖİR Rapor No
        26: None,            # Z: TSE ÜGM ÖİR Rapor Tarihi
        27: None,            # AA: İptal Nedeni
        28: None,            # AB: İptal Nedeni Açıklama
        29: None,            # AC: Vardiya Sayısı
    }

    border = make_thin_border()
    wrap_align = Alignment(wrap_text=True, vertical='top')
    center_align = Alignment(wrap_text=True, vertical='top', horizontal='center')
    bold_font = Font(bold=True, size=11)
    normal_font = Font(size=10)

    for col, value in data.items():
        cell = ws.cell(row=row_num, column=col)
        cell.value = value
        cell.border = border
        if col in [2, 3, 4, 5]:
            cell.font = bold_font
            cell.alignment = wrap_align
        elif col in [16, 17, 18, 19, 20]:
            cell.font = normal_font
            cell.alignment = center_align
        else:
            cell.font = normal_font
            cell.alignment = wrap_align


def write_empty_month_row(ws, row_num, ay, yil):
    """Veri olmayan ay için boş satır"""
    border = make_thin_border()
    bold_font = Font(bold=True, size=11)
    normal_font = Font(size=10)
    wrap_align = Alignment(wrap_text=True, vertical='top')

    for col in range(1, 30):
        cell = ws.cell(row=row_num, column=col)
        cell.border = border
        cell.font = normal_font
        cell.alignment = wrap_align

    ws.cell(row=row_num, column=1).value = 'DENİZLİ'
    ws.cell(row=row_num, column=2).value = ay
    ws.cell(row=row_num, column=2).font = bold_font
    ws.cell(row=row_num, column=3).value = yil
    ws.cell(row=row_num, column=3).font = bold_font
    # İzahat kural 5: başvuru yoksa muafiyet tarihi sütununa belirt
    ws.cell(row=row_num, column=6).value = 'Muafiyet Başvurusu bulunmamaktadır'
    ws.cell(row=row_num, column=6).font = Font(size=10, italic=True)


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['2026-9 ve 32 ÜGM']

    # Mevcut veri satırlarını temizle (5-16)
    for row in range(5, 17):
        for col in range(1, 30):
            ws.cell(row=row, column=col).value = None

    current_row = 5

    # Her başvurudaki her nihai ürün + girdi için satır yaz
    for basvuru in BASVU_DATA:
        for nihai_urun_adi, girdi_listesi in basvuru['nihai_urunler']:
            for girdi in girdi_listesi:
                write_data_row(ws, current_row, basvuru['ay'], basvuru['yil'],
                               basvuru, nihai_urun_adi, girdi)
                ws.row_dimensions[current_row].height = 30
                current_row += 1

    # Kalan aylar için boş satırlar
    for ay in REMAINING_MONTHS:
        write_empty_month_row(ws, current_row, ay, 2026)
        ws.row_dimensions[current_row].height = 25
        current_row += 1

    wb.save(EXCEL_PATH)
    print(f'Excel kaydedildi. Toplam veri satırı: {current_row - 5}')
    print(f'Veri satır sayısı (başvurular): {current_row - 5 - len(REMAINING_MONTHS)}')


if __name__ == '__main__':
    main()
